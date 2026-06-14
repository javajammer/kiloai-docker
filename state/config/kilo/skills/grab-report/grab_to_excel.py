import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

def parse_grab_data(raw_data):
    """Parse raw Grab data into structured records."""
    lines = raw_data.strip().split('\n')
    records = []
    
    for line in lines[1:]:
        parts = line.split('\t')
        if len(parts) < 6:
            continue
        
        tanggal_jam = parts[0]
        tipe_layanan = parts[4]
        
        # Handle inconsistent column count (some rows have extra "Bandung" column)
        jumlah = None
        
        if len(parts) >= 8 and parts[5] == "Bandung":
            jumlah = parts[7]
        elif len(parts) >= 7:
            jumlah = parts[6]
        
        if jumlah is None:
            continue
        
        tanggal = tanggal_jam.split(',')[0].strip()
        
        pickup = parts[2].strip()
        dropoff = parts[3].strip()
        
        # Normalize pickup location
        if pickup in ["Cottonwood Bed Breakfast House - Sukawarna", "Sukawarna"]:
            pickup = "Cottonwood (Hotel)"
        elif pickup == "Zoe Guest House Hotel Bandung":
            pickup = "Zoe (Hotel)"
        elif pickup == "Near Block B4 No.26 Cluster Green Hill Serpong Garden 2 Residence":
            pickup = "Serpong Garden 2 Cluster Green Hill (Rumah)"
        elif pickup == "No.25E Jalan Jaya Mandala IV":
            pickup = "Jalan Jaya Mandala IV (Jakarta)"
        elif pickup in ["Infokes Ind", "Mitrais"]:
            pickup = "Kantor Infokes Bandung"
        
        # Normalize dropoff location
        if dropoff in ["Cottonwood Bed Breakfast House - Sukawarna", "Sukawarna"]:
            dropoff = "Cottonwood (Hotel)"
        elif dropoff == "Zoe Guest House Hotel Bandung":
            dropoff = "Zoe (Hotel)"
        elif dropoff == "Near Block B4 No.26 Cluster Green Hill Serpong Garden 2 Residence":
            dropoff = "Serpong Garden 2 Cluster Green Hill (Rumah)"
        elif dropoff == "No.25E Jalan Jaya Mandala IV":
            dropoff = "Jalan Jaya Mandala IV (Jakarta)"
        elif dropoff in ["Infokes Ind", "Mitrais"]:
            dropoff = "Kantor Infokes Bandung"
        
        lokasi = f"{pickup} - {dropoff}"
        
        # Normalize service type
        jenis = tipe_layanan
        if jenis in ["Car Standard", "GrabCar", "GrabCar Hemat", "GrabCar XL"]:
            jenis = "GrabCar"
        elif jenis in ["Bike Standard", "GrabBike", "GrabBike Hemat", "GrabBike XL"]:
            jenis = "GrabBike"
        
        amount = float(jumlah)
        
        records.append({
            'TANGGAL': tanggal,
            'LOKASI': lokasi,
            'Jenis': jenis,
            'KM': '',
            'BENSIN': '',
            'Transport': amount,
            'Penginapan': '',
            'UPD/Uang Makan': '',
            'Deskripsi': '',
            'Lain-lain': '',
            'TOTAL': amount
        })
    
    return records

def create_excel(records, output_path='/tmp/Business_Expense.xlsx'):
    """Create formatted Excel file from records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Expense"
    
    headers = ['TANGGAL', 'LOKASI', 'Jenis', 'KM', 'BENSIN', 'Transport', 
               'Penginapan', 'UPD/Uang Makan', 'Deskripsi', 'Lain-lain', 'TOTAL']
    
    # Write headers with bold font
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Write data rows
    prev_date = None
    for idx, row in enumerate(records, 2):
        tanggal = '' if row['TANGGAL'] == prev_date else row['TANGGAL']
        prev_date = row['TANGGAL']
        
        ws.cell(row=idx, column=1, value=tanggal)
        ws.cell(row=idx, column=2, value=row['LOKASI'])
        ws.cell(row=idx, column=3, value=row['Jenis'])
        ws.cell(row=idx, column=4, value=row['KM'])
        ws.cell(row=idx, column=5, value=row['BENSIN'])
        ws.cell(row=idx, column=6, value=row['Transport'])
        ws.cell(row=idx, column=7, value=row['Penginapan'])
        ws.cell(row=idx, column=8, value=row['UPD/Uang Makan'])
        ws.cell(row=idx, column=9, value=row['Deskripsi'])
        ws.cell(row=idx, column=10, value=row['Lain-lain'])
        ws.cell(row=idx, column=11, value=row['TOTAL'])
    
    # Add Total row
    total_row = len(records) + 2
    ws.cell(row=total_row, column=5, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})").font = Font(bold=True)
    ws.cell(row=total_row, column=11, value=f"=SUM(K2:K{total_row-1})").font = Font(bold=True)
    
    # Add Total Pengeluaran row
    total_pengeluaran_row = total_row + 1
    ws.cell(row=total_pengeluaran_row, column=5, value="Total Pengeluaran").font = Font(bold=True)
    ws.cell(row=total_pengeluaran_row, column=6, value=f"=F{total_row}").font = Font(bold=True)
    ws.cell(row=total_pengeluaran_row, column=11, value=f"=K{total_row}").font = Font(bold=True)
    
    # Apply currency format to Transport and TOTAL columns
    for col in [6, 11]:
        for row in range(2, total_row + 1):
            ws.cell(row=row, column=col).number_format = '_(* #,##0_);_(* (#,##0);_(*  - ??_);_(@_)'
    
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        with open(sys.argv[1], 'r') as f:
            raw_data = f.read()
    elif not sys.stdin.isatty():
        raw_data = sys.stdin.read()
    else:
        # Default sample data - user should replace with actual data
        print("Usage: python grab_to_excel.py <input_file>")
        print("Or pipe data via stdin")
        sys.exit(1)
    
    records = parse_grab_data(raw_data)
    create_excel(records)
    print(f"File Business_Expense.xlsx berhasil dibuat dengan {len(records)} transaksi")